import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as mbox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
import win32com.client as win32
import os
import shutil

# set the directory to a local folder in user's computer
DESKTOP_DIR = Path.home() / "OneDrive" / "Desktop" / "Invoice" / "details report"
DESKTOP_DIR.mkdir(parents=True, exist_ok=True) # create folder accordingly if the folders are not found

#
font_bold = Font(bold=True)
center_middle_alignment = Alignment(horizontal='center', vertical='center')

def select_folders(title):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(initialdir=str(DESKTOP_DIR), title=title)
    return folder_path

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[cell.column_letter].width = adjusted_width

def consolidate_txt_to_excel(folder_path):
    try:
        detailsReportPath = DESKTOP_DIR / 'details.xlsx'
        wb = openpyxl.load_workbook(detailsReportPath)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Invoice'
    else:
        # If "Invoice" sheet exists, delete it
        if "Invoice" in wb.sheetnames:
            del wb["Invoice"]
        # Create a new "Invoice" sheet
        ws = wb.create_sheet("Invoice")
       
    headers = ["FILE NAME #", "BATCH #", "INVOICE #", "VALUE (MYR)"]
    ws.append(headers)

    # Define patterns for extracting required details from the text file
    transfer_num_pattern = re.compile(r"Transfer Num:\s*(\d+)")
    jda_batch_pattern = re.compile(r"JDA Batch Number:\s*(\d+)")
    invoice_number_pattern = re.compile(r"Invoice Number:\s*(\d+)")
    value_pattern = re.compile(r"(\d+\.\d{2})(?=[^\d]*?$|(?=[^\d]*?))")

    for file_path in Path(folder_path).rglob('*.txt'):
        with open(file_path, 'r') as file:
            content = file.read()
            # Extract details using patterns
            transfer_num = transfer_num_pattern.search(content)
            jda_batch = jda_batch_pattern.search(content)
            invoice_number = invoice_number_pattern.search(content)
            value = value_pattern.search(content)
            row_data = [
                f"{transfer_num.group(1)}.TXT" if transfer_num else None,
                jda_batch.group(1) if jda_batch else None,
                invoice_number.group(1) if invoice_number else None,
                value.group(1) if value else None
            ]
            ws.append(row_data)

    autofit_columns(ws)
    for cell in ws["A1:Z1"][0]:
        cell.font = font_bold
        cell.alignment = center_middle_alignment
    
    remove_duplicates(ws)
    wb.save(DESKTOP_DIR / 'details.xlsx')

def remove_duplicates(ws):
    unique_rows = set()
    rows_to_delete = []

    for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming row 1 has headers, so we start from row 2
        # Convert the row into a tuple to make it hashable
        row_data = tuple(row)
        if row_data in unique_rows:
            rows_to_delete.append(row[0].row)  # Save the row number to delete later
        else:
            unique_rows.add(row_data)

    # Deleting rows needs to be done from bottom to top to avoid changing indices
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)


def convert_xml_to_xlsx(folder_path):
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
    except AttributeError:
        # Path to the gen_py directory
        gen_py_path = os.path.expanduser('~\\AppData\\Local\\Temp\\gen_py')
        
        # Check if the directory exists
        if os.path.exists(gen_py_path):
            print('Deleting ...', gen_py_path)
            # Attempt to delete the directory
            shutil.rmtree(gen_py_path)

        # Retry dispatching Excel Application
        excel = win32.gencache.EnsureDispatch('Excel.Application')

    initial_workbooks = set(excel.Workbooks)
    
    for file_path in Path(folder_path).rglob('*.xls'):
        # Check if the corresponding .xlsx file already exists to avoid re-conversion
        if not Path(str(file_path) + 'x').exists():
            wb = excel.Workbooks.Open(str(file_path))
            wb.SaveAs(str(file_path) + 'x', FileFormat=51)  # 51 represents xlsx format
            wb.Close()
            os.remove(file_path) # delete original '.xls' reports

    # Don't close Excel if there were other workbooks open initially
    if not set(excel.Workbooks) - initial_workbooks:
        excel.Application.Quit()


# read through each excel files and gather the rows
def Inv_Report(folder_path):
    extracted_dataframes = []
    
    # Iterate through all files in the specified folder
    for file in Path(folder_path).rglob('*.xlsx'):
        print(file, '.xlsx found')
        df = pd.read_excel(file, engine='openpyxl')
        numeric_rows = df[df.iloc[:, 2].astype(str).str.isnumeric()]
        extracted_dataframes.append(numeric_rows)
        
    if extracted_dataframes:
        combined_df = pd.concat(extracted_dataframes, ignore_index=True)
        combined_df.columns = ["Order", "JDA #", "SKU", "Color", "Size", "Style", "Description", "Department  Code", 
               "UOM", "Total Units", "Units Cost", "Ext.   Cost", "Coutry OfOrigin", "HS Code"]
        combined_df = combined_df.drop(columns=["Units Cost", "Ext.   Cost", "Coutry OfOrigin"],errors='ignore')
        return combined_df
    else:
        print("No dataframes available for concatenation.")
        return None

    
def add_to_details_report(details_report_path, all_rows):
    if all_rows is None or all_rows.empty:
        return

    # Check if file exists, load it if it does. Otherwise, create a new workbook
    try:
        wb = openpyxl.load_workbook(details_report_path)
        print(details_report_path, " found")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'SKU'
    else:
        # If "Invoice" sheet exists, delete it
        if "SKU" in wb.sheetnames:
            del wb["SKU"]
            print("sheet SKU found")
        # Create a new "Invoice" sheet
        ws = wb.create_sheet("SKU")

    # Add headers to the "Invoice" sheet
    headers = ["Order", "JDA #", "SKU", "Color", "Size", "Style", "Description", "Department  Code", 
               "UOM", "Total Units", "HS Code"]
    ws.append(headers)

    for _, row in all_rows.iterrows():
        ws.append(list(row))

    autofit_columns(ws)
    for cell in ws["A1:Z1"][0]:
        cell.font = font_bold
        cell.alignment = center_middle_alignment
    wb.save(details_report_path)

    
def reorganize_invoice_report():
    selected_folder = select_folders("Select Invoice Report folder")
    if not selected_folder:
        return

    # Convert .xls files to .xlsx in the selected folder
    convert_xml_to_xlsx(selected_folder)
    extracted_rows = Inv_Report(selected_folder)
    if extracted_rows is not None:
        add_to_details_report(DESKTOP_DIR / 'details.xlsx', extracted_rows)
        mbox.showinfo("Info", f"Reorganization completed and added to {DESKTOP_DIR / 'details.xlsx'}")


def main():
    root = tk.Tk()
    root.title("Generate Details Report")
    root.geometry('300x200')

    def exit_app():
        root.quit()
        root.destroy()

    def get_choice(val):
        if val == "1":
            DESKTOP_DIR = Path.home() / "OneDrive" / "Desktop" / "Invoice" / "details report"
            DESKTOP_DIR.mkdir(parents=True, exist_ok=True)

            selected_folder = select_folders("Select folder containing TXT Invoices")
            if selected_folder:                
                consolidate_txt_to_excel(selected_folder)
                mbox.showinfo("Success", f"Consolidation completed and saved to {DESKTOP_DIR / 'details.xlsx'}")

        elif val == "2":
            DESKTOP_DIR = Path.home() / "OneDrive" / "Desktop" / "Invoice" / "details report"
            DESKTOP_DIR.mkdir(parents=True, exist_ok=True)
        
            reorganize_invoice_report()
        else:
            exit_app()

    lbl = tk.Label(root, text="Please select the operation you'd like to execute:")
    lbl.pack(pady=12, padx=10)

    btn_txt_invoices = tk.Button(root, text="Consolidate TXT Invoices", command=lambda: get_choice("1"))
    btn_txt_invoices.pack(pady=12)

    btn_report = tk.Button(root, text="Reorganize Invoice Report", command=lambda: get_choice("2"))
    btn_report.pack(pady=12)

    root.mainloop()

if __name__ == "__main__":
    main()
